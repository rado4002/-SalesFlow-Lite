from __future__ import annotations

import io
import csv
from typing import Any, Dict, List

from fastapi import UploadFile
from openpyxl import load_workbook

from src.clients.java_products_client import JavaProductsClient


# --------------------------------------------------
# Helpers
# --------------------------------------------------

def _norm(s: str) -> str:
    return s.strip().lower().replace(" ", "_").replace("-", "_")


def _to_int(v: Any) -> int | None:
    try:
        return int(float(v))
    except Exception:
        return None


def _to_float(v: Any) -> float | None:
    try:
        return float(v)
    except Exception:
        return None


# --------------------------------------------------
# Readers (CSV / Excel)
# --------------------------------------------------

async def _read_file(file: UploadFile) -> List[Dict[str, Any]]:
    content = await file.read()
    if not content:
        raise ValueError("Empty file")

    filename = (file.filename or "").lower()

    # ---------- CSV ----------
    if filename.endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [{_norm(k): v for k, v in row.items()} for row in reader]

    # ---------- XLSX ----------
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    header = [_norm(str(c.value)) for c in ws[1]]
    rows: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        obj = dict(zip(header, row))
        if any(v not in (None, "") for v in obj.values()):
            rows.append(obj)

    return rows


# --------------------------------------------------
# Canonical Sales Processor
# --------------------------------------------------

async def process_sales_upload_to_canonical(
    file: UploadFile,
    token: str,
) -> Dict[str, Any]:
    """
    Converts Excel/CSV sales file into canonical Java format.

    Output items:
        { productId, sku, quantity }
    """

    rows = await _read_file(file)

    if not rows:
        return {
            "items": [],
            "errors": [],
            "total_rows": 0,
            "valid_rows": 0,
        }

    # --------------------------------------------------
    # JOIN with Java products
    # --------------------------------------------------
    products_client = JavaProductsClient(token)
    try:
        products = await products_client.get_all_products()
    finally:
        await products_client.close()

    by_id = {p.id: p for p in products if p.id is not None}
    by_sku = {p.sku.lower(): p for p in products if p.sku}
    by_name = {p.name.lower(): p for p in products if p.name}

    items: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for row_index, row in enumerate(rows, start=2):
        # --------------------------------------------------
        # One-of rule: product_id | sku | name
        # --------------------------------------------------
        if not row.get("product_id") and not row.get("sku") and not row.get("name"):
            errors.append({
                "row": row_index,
                "field": "product_id / sku / name",
                "reason": "One of product_id, sku or name is required",
            })
            continue

        # --------------------------------------------------
        # Quantity validation
        # --------------------------------------------------
        quantity = _to_float(row.get("quantity"))
        if quantity is None or quantity <= 0:
            errors.append({
                "row": row_index,
                "field": "quantity",
                "reason": "Invalid or <= 0",
            })
            continue

        # --------------------------------------------------
        # Resolve product
        # --------------------------------------------------
        product = None

        pid = _to_int(row.get("product_id"))
        if pid and pid in by_id:
            product = by_id[pid]
        elif row.get("sku"):
            product = by_sku.get(str(row["sku"]).lower())
        elif row.get("name"):
            product = by_name.get(str(row["name"]).lower())

        if not product:
            errors.append({
                "row": row_index,
                "field": "product",
                "reason": "Unknown product",
            })
            continue

        # --------------------------------------------------
        # Canonical output (Java SaleItemRequest)
        # --------------------------------------------------
        items.append({
            "productId": product.id,
            "sku": product.sku,
            "quantity": int(quantity),
        })

    return {
        "items": items,
        "errors": errors,
        "total_rows": len(rows),
        "valid_rows": len(items),
    }
