# src/services/report_service.py
from __future__ import annotations

import io
from datetime import datetime
from typing import Literal, Tuple, Optional,Union
from pathlib import Path

from src.models.report import ReportRequest

from fastapi import HTTPException
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.barcharts import VerticalBarChart

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import LineChart, Reference

from src.models.dto.analytics_dto import (
    AnalyticsPeriod,
    SalesAnalyticsResponse,
    StockAnalyticsResponse,
)
from src.services.analytics_service import (
    compute_sales_analytics,
    compute_stock_analytics,
)

ReportFormat = Literal["pdf", "excel"]
ReportType = Literal["sales", "stock", "combined"]


# =====================================================================
# INTERNAL HELPERS — COMMON
# =====================================================================

def _now_label() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def _period_label(period: AnalyticsPeriod) -> str:
    return period.value


# =====================================================================
# PDF GENERATION — EXECUTIVE / INVESTOR STYLE
# =====================================================================

def _build_cover_page(
    styles,
    title: str,
    period: AnalyticsPeriod,
) -> list:
    """
    Page de garde “Executive / Investor report”.
    Grande marge, titre centré, sous-titre, méta d’horodatage.
    """
    elements: list = []

    # On repart de Title mais on renforce les paramètres
    cover_title = ParagraphStyle(
        "CoverTitle",
        parent=styles["Title"],
        fontSize=28,
        leading=32,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
    )

    cover_subtitle = ParagraphStyle(
        "CoverSubtitle",
        parent=styles["Normal"],
        fontSize=13,
        leading=18,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#4b5563"),
    )

    meta_style = ParagraphStyle(
        "CoverMeta",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#6b7280"),
    )

    footer_style = ParagraphStyle(
        "CoverFooter",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#9ca3af"),
    )

    # Gros espace pour centrer verticalement
    elements.append(Spacer(1, 5 * cm))
    elements.append(
        Paragraph("SalesFlow Lite", ParagraphStyle(
            "Brand",
            parent=styles["Normal"],
            fontSize=12,
            textColor=colors.HexColor("#2563eb"),
            alignment=TA_CENTER,
            leading=16,
        ))
    )
    elements.append(Spacer(1, 0.5 * cm))

    elements.append(Paragraph(title, cover_title))
    elements.append(Spacer(1, 0.4 * cm))

    subtitle_text = (
        "Executive analytics report consolidating revenue, margins and inventory health. "
        "Designed for decision-makers, investors and finance leaders."
    )
    elements.append(Paragraph(subtitle_text, cover_subtitle))
    elements.append(Spacer(1, 1.5 * cm))

    meta = (
        f"Period analysed: <b>{period.value}</b><br/>"
        f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    elements.append(Paragraph(meta, meta_style))

    # Bandeau discret en bas de page
    elements.append(Spacer(1, 7 * cm))
    elements.append(
        Paragraph(
            "SalesFlow Lite • CSDE Portfolio Project • Confidential analytics snapshot",
            footer_style,
        )
    )

    return elements


def _build_branding_header(styles, section_title: str, period: AnalyticsPeriod) -> list:
    """
    Header en haut de la première page de contenu (après cover).
    Plus discret que la cover.
    """
    elems: list = []

    h_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading1"],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#111827"),
    )
    meta_style = ParagraphStyle(
        "SectionMeta",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#6b7280"),
    )

    elems.append(Paragraph(section_title, h_style))
    elems.append(Spacer(1, 0.15 * cm))

    meta = (
        f"Period: <b>{period.value}</b> • "
        f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')} • "
        "Single source of truth (same logic as dashboard)"
    )
    elems.append(Paragraph(meta, meta_style))
    elems.append(Spacer(1, 0.4 * cm))

    return elems


def _build_executive_summary(
    styles,
    sales: SalesAnalyticsResponse,
    stock: Optional[StockAnalyticsResponse],
) -> list:
    """
    Bloc “Executive summary” avec 3–5 bullet points lisibles pour un investisseur.
    """
    elems: list = []

    title_style = ParagraphStyle(
        "ExecTitle",
        parent=styles["Heading2"],
        fontSize=15,
        textColor=colors.HexColor("#111827"),
    )
    bullet_style = ParagraphStyle(
        "ExecBullet",
        parent=styles["Normal"],
        fontSize=10.5,
        leading=14,
        textColor=colors.HexColor("#374151"),
        leftIndent=8,
        bulletIndent=0,
    )

    k = sales.kpis
    total_rev = f"{k.total_revenue:,.0f}"
    avg_ticket = f"{k.average_ticket:,.0f}"
    tx = f"{k.total_transactions:,.0f}"
    trend = k.seasonal_hint or "no clear seasonal pattern"

    elems.append(Paragraph("Executive summary", title_style))
    elems.append(Spacer(1, 0.15 * cm))

    bullet_points = [
        f"Total revenue for the period reached <b>{total_rev}</b> across <b>{tx}</b> transactions.",
        f"Average ticket is <b>{avg_ticket}</b>, indicating the typical basket size for active customers.",
        f"Current trend shows <b>{trend}</b> based on recent sales behaviour.",
    ]

    if stock:
        ks = stock.kpis
        low_ratio = f"{ks.low_stock_ratio:,.1f}%"
        urgent = ks.urgent_reorder_count
        dead = ks.dead_stock_count
        bullet_points.append(
            f"Inventory health: <b>{low_ratio}</b> of SKUs are in low-stock range, with "
            f"<b>{urgent}</b> items below 7 days of coverage and <b>{dead}</b> dead-stock references."
        )

    if k.top_products:
        main = k.top_products[0]
        bullet_points.append(
            f"Top performing product: <b>{main.name}</b> accounts for "
            f"<b>{main.share_of_revenue:,.1f}%</b> of revenue in this period."
        )

    for text in bullet_points:
        elems.append(Paragraph(f"• {text}", bullet_style))
        elems.append(Spacer(1, 0.05 * cm))

    elems.append(Spacer(1, 0.4 * cm))
    return elems


def _build_sales_kpi_table(sales: SalesAnalyticsResponse) -> Table:
    k = sales.kpis
    data = [
        ["Metric", "Value"],
        ["Total revenue", f"{k.total_revenue:,.2f}"],
        ["Total quantity", f"{k.total_quantity:,.0f}"],
        ["Transactions", f"{k.total_transactions:,.0f}"],
        ["Average ticket", f"{k.average_ticket:,.2f}"],
        ["Trend", k.seasonal_hint or "N/A"],
    ]
    table = Table(data, hAlign="LEFT", colWidths=[5 * cm, 5 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _build_stock_kpi_table(stock: StockAnalyticsResponse) -> Table:
    k = stock.kpis
    data = [
        ["Metric", "Value"],
        ["Total stock value", f"{k.total_stock_value:,.2f}"],
        ["Out of stock products", f"{k.out_of_stock_count}"],
        ["Low stock products", f"{k.low_stock_count}"],
        ["Low stock ratio (%)", f"{k.low_stock_ratio:,.2f}"],
        ["Rotation / year", f"{k.rotation_per_year:,.2f}" if k.rotation_per_year else "N/A"],
        [
            "Average coverage (days)",
            f"{k.avg_coverage_days:,.1f}" if k.avg_coverage_days is not None else "N/A",
        ],
        ["Urgent reorder (coverage < 7d)", f"{k.urgent_reorder_count}"],
        ["Dead stock", f"{k.dead_stock_count}"],
    ]
    table = Table(data, hAlign="LEFT", colWidths=[7 * cm, 4 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    return table


def _build_top_products_table(sales: SalesAnalyticsResponse) -> Optional[Table]:
    top = sales.kpis.top_products or []
    if not top:
        return None

    data = [["Product", "Qty", "Revenue", "Share of revenue (%)"]]
    for p in top:
        data.append(
            [
                p.name,
                f"{p.total_quantity:,.0f}",
                f"{p.revenue:,.2f}",
                f"{p.share_of_revenue:,.1f}",
            ]
        )

    table = Table(data, hAlign="LEFT", colWidths=[7 * cm, 2 * cm, 3 * cm, 3 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _build_daily_sales_chart(sales: SalesAnalyticsResponse) -> Optional[Drawing]:
    if not sales.daily:
        return None

    dates = [dp.date for dp in sales.daily]
    revenues = [dp.total_revenue for dp in sales.daily]

    x_data = list(range(len(dates)))
    data = list(zip(x_data, revenues))

    drawing = Drawing(400, 180)
    lp = LinePlot()
    lp.x = 50
    lp.y = 30
    lp.width = 320
    lp.height = 120
    lp.data = [data]
    lp.joinedLines = True
    lp.lines[0].strokeColor = colors.HexColor("#2563eb")
    lp.lineLabelFormat = None

    lp.xValueAxis.valueMin = 0
    lp.xValueAxis.valueMax = max(x_data) if x_data else 1
    lp.xValueAxis.valueStep = max(1, len(x_data) // 6)

    lp.yValueAxis.valueMin = 0
    lp.yValueAxis.valueMax = max(revenues) * 1.1 if revenues else 1
    lp.yValueAxis.valueStep = max(1, int(lp.yValueAxis.valueMax / 4))

    drawing.add(lp)
    return drawing


def _build_stock_status_chart(stock: StockAnalyticsResponse) -> Optional[Drawing]:
    snaps = stock.critical_products or []
    if not snaps:
        return None

    status_counts = {"OK": 0, "LOW": 0, "OUT": 0, "DEAD": 0}
    for s in snaps:
        name = (s.status.name or "").upper()
        if "DEAD" in name:
            status_counts["DEAD"] += 1
        elif "OUT" in name:
            status_counts["OUT"] += 1
        elif "LOW" in name:
            status_counts["LOW"] += 1
        else:
            status_counts["OK"] += 1

    labels = ["OK", "LOW", "OUT", "DEAD"]
    values = [status_counts[k] for k in labels]

    drawing = Drawing(400, 180)
    bc = VerticalBarChart()
    bc.x = 50
    bc.y = 30
    bc.width = 320
    bc.height = 120
    bc.data = [values]
    bc.categoryAxis.categoryNames = labels
    bc.barWidth = 12
    bc.strokeColor = colors.transparent
    bc.bars[0].fillColor = colors.HexColor("#10b981")

    drawing.add(bc)
    return drawing

def _auto_adjust_column_width(ws):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0

        for cell in column_cells:
            if isinstance(cell, MergedCell):
                continue

            if cell.value is not None:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)





def _generate_sales_pdf(
    sales: SalesAnalyticsResponse,
    stock: Optional[StockAnalyticsResponse] = None,
) -> bytes:
    """
    Génère un PDF multi-page “Executive / Investor report” :
      - Page de garde
      - Executive summary
      - Sales KPIs + courbe
      - Top products
      - Inventory & stock health (si fourni)
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    body_style = styles["Normal"]
    body_style.fontSize = 10
    body_style.leading = 14

    elements: list = []

    # ------------------- COVER PAGE -------------------
    main_title = (
        "Sales Analytics Report"
        if stock is None
        else "Sales & Stock Analytics Report"
    )
    elements += _build_cover_page(styles, main_title, sales.period)
    elements.append(PageBreak())

    # ------------------- HEADER + EXEC SUMMARY -------------------
    elements += _build_branding_header(
        styles,
        "Executive analytics overview",
        sales.period,
    )
    elements += _build_executive_summary(styles, sales, stock)

    intro = (
        "The following pages provide a consolidated view of sales performance, "
        "inventory exposure and product dynamics for the selected period. "
        "All indicators are computed from the analytics service, using the "
        "same logic as the operational dashboard, ensuring consistency and "
        "auditability for leadership and investors."
    )
    elements.append(Paragraph(intro, body_style))
    elements.append(Spacer(1, 0.6 * cm))

    # ------------------- SALES KPI SECTION -------------------
    elements.append(Paragraph("<b>1. Sales KPIs</b>", styles["Heading2"]))
    elements.append(Spacer(1, 0.2 * cm))

    elements.append(_build_sales_kpi_table(sales))
    elements.append(Spacer(1, 0.4 * cm))

    chart = _build_daily_sales_chart(sales)
    if chart:
        elements.append(Paragraph("Daily revenue timeline", styles["Heading3"]))
        elements.append(Spacer(1, 0.2 * cm))
        elements.append(chart)
        elements.append(Spacer(1, 0.6 * cm))

    # ------------------- TOP PRODUCTS -------------------
    elements.append(Paragraph("<b>2. Top products</b>", styles["Heading2"]))
    elements.append(Spacer(1, 0.2 * cm))

    top_table = _build_top_products_table(sales)
    if top_table:
        elements.append(top_table)
    else:
        elements.append(Paragraph("No product sales available for this period.", body_style))

    elements.append(Spacer(1, 0.6 * cm))

    # ------------------- STOCK SECTION (OPTIONAL) -------------------
    if stock:
        elements.append(PageBreak())
        elements += _build_branding_header(
            styles, "Inventory & stock risk", sales.period
        )

        elements.append(Paragraph("<b>3. Inventory KPIs</b>", styles["Heading2"]))
        elements.append(Spacer(1, 0.2 * cm))

        elements.append(_build_stock_kpi_table(stock))
        elements.append(Spacer(1, 0.4 * cm))

        s_chart = _build_stock_status_chart(stock)
        if s_chart:
            elements.append(Paragraph("Critical stock distribution", styles["Heading3"]))
            elements.append(Spacer(1, 0.2 * cm))
            elements.append(s_chart)
            elements.append(Spacer(1, 0.5 * cm))

        commentary = (
            "Inventory KPIs highlight where working capital is locked in low-moving or "
            "dead stock, and where service risk is high due to low coverage. "
            "Combining these indicators with sales velocity makes it possible to "
            "prioritize replenishment, renegotiate with suppliers, or liquidate "
            "non-performing items."
        )
        elements.append(Paragraph(commentary, body_style))

    # ------------------- FOOTER -------------------
    footer_style = ParagraphStyle(
        "footer",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#6b7280"),
        alignment=TA_CENTER,
    )
    elements.append(Spacer(1, 0.8 * cm))
    elements.append(
        Paragraph(
            "SalesFlow Lite – CSDE Project • Auto-generated analytics report for leadership and investors.",
            footer_style,
        )
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


# =====================================================================
# EXCEL GENERATION (inchangé, déjà “pro”)
# =====================================================================

def _generate_sales_excel(
    sales: SalesAnalyticsResponse,
    stock: Optional[StockAnalyticsResponse] = None,
) -> bytes:
    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "Overview"

    ws_overview["A1"] = "SalesFlow Lite – Analytics Report"
    ws_overview["A1"].font = Font(size=16, bold=True)
    ws_overview.merge_cells("A1:D1")

    ws_overview["A2"] = "Generated at"
    ws_overview["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_overview["A3"] = "Period"
    ws_overview["B3"] = sales.period.value

    k = sales.kpis
    ws_overview["A5"] = "Sales KPIs"
    ws_overview["A5"].font = Font(bold=True)

    rows = [
        ("Total revenue", k.total_revenue),
        ("Total quantity", k.total_quantity),
        ("Transactions", k.total_transactions),
        ("Average ticket", k.average_ticket),
        ("Trend", k.seasonal_hint or "N/A"),
    ]
    row_idx = 6
    for label, value in rows:
        ws_overview[f"A{row_idx}"] = label
        ws_overview[f"B{row_idx}"] = value
        row_idx += 1

    if stock:
        ks = stock.kpis
        row_idx += 1
        ws_overview[f"A{row_idx}"] = "Stock KPIs"
        ws_overview[f"A{row_idx}"].font = Font(bold=True)
        row_idx += 1
        stock_rows = [
            ("Total stock value", ks.total_stock_value),
            ("Out of stock", ks.out_of_stock_count),
            ("Low stock", ks.low_stock_count),
            ("Low stock ratio (%)", ks.low_stock_ratio),
            ("Rotation / year", ks.rotation_per_year or 0),
            ("Avg coverage (days)", ks.avg_coverage_days or 0),
            ("Urgent reorder (<7d)", ks.urgent_reorder_count),
            ("Dead stock", ks.dead_stock_count),
        ]
        for label, value in stock_rows:
            ws_overview[f"A{row_idx}"] = label
            ws_overview[f"B{row_idx}"] = value
            row_idx += 1

    ws_daily = wb.create_sheet("Daily Sales")
    ws_daily.append(["Date", "Revenue", "Quantity", "Transactions"])
    for cell in ws_daily[1]:
        cell.font = Font(bold=True)

    for dp in sales.daily:
        ws_daily.append([
            dp.date.isoformat() if hasattr(dp.date, "isoformat") else str(dp.date),
            float(dp.total_revenue),
            float(dp.total_quantity),
            int(dp.total_transactions),
        ])

    if sales.daily:
        chart = LineChart()
        chart.title = "Daily revenue"
        chart.y_axis.title = "Revenue"
        chart.x_axis.title = "Day index"

        data_ref = Reference(
            ws_daily,
            min_col=2,
            min_row=1,
            max_row=1 + len(sales.daily),
            max_col=2,
        )
        chart.add_data(data_ref, titles_from_data=True)
        ws_daily.add_chart(chart, "F2")

    ws_top = wb.create_sheet("Top Products")
    ws_top.append(["Product", "Quantity", "Revenue", "Share (%)"])
    for cell in ws_top[1]:
        cell.font = Font(bold=True)

    for p in sales.kpis.top_products or []:
        ws_top.append([
            p.name,
            float(p.total_quantity),
            float(p.revenue),
            float(p.share_of_revenue),
        ])

    if stock:
        ws_stock = wb.create_sheet("Critical Stock")
        ws_stock.append([
            "Product",
            "SKU",
            "Current stock",
            "Stock value",
            "Coverage (days)",
            "Status",
            "Last sale date",
        ])
        for cell in ws_stock[1]:
            cell.font = Font(bold=True)

        for s in stock.critical_products or []:
            ws_stock.append([
                s.product_id,
                s.name,
                float(s.current_stock),
                float(s.stock_value),
                float(s.coverage_days) if s.coverage_days is not None else None,
                s.status.name if s.status else None,
                s.last_sale_date.isoformat() if s.last_sale_date else None,
            ])

    # ✅ SAFE auto-width
    for ws in wb.worksheets:
        _auto_adjust_column_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()



# =====================================================================
# PUBLIC API
# =====================================================================
REPORTS_DIR = Path("reports/generated")


def _save_report_to_disk(content: bytes, filename: str) -> str:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    file_path = REPORTS_DIR / filename
    file_path.write_bytes(content)

    return str(file_path)

def get_latest_report(
    report_type: ReportType,
    fmt: ReportFormat,
) -> Path:
    pattern = f"analytics_{report_type}_*_.{ 'pdf' if fmt == 'pdf' else 'xlsx' }"

    files = sorted(
        REPORTS_DIR.glob(pattern),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    if not files:
        raise FileNotFoundError("No generated report found")

    return files[0]

def find_last_report_on_disk(
    report_type: str,
    fmt: str,
) -> Optional[Path]:
    if not REPORTS_DIR.exists():
        return None

    files = sorted(
        REPORTS_DIR.glob(f"analytics_{report_type}_*.{fmt}"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )

    return files[0] if files else None



async def generate_report(
    report_type: ReportType,
    fmt: ReportFormat,
    period: AnalyticsPeriod,
    token: Optional[str],
    save_to_disk: bool = False,
) -> Union[Tuple[bytes, str, str], dict]:
    """
    Core report generator.

    Modes:
    - save_to_disk = False → API download (bytes)
    - save_to_disk = True  → Scheduler (disk)

    Returns:
    - API mode  → (content, filename, media_type)
    - Disk mode → { file_path, filename, generated_at }
    """

    # --------------------------------------------------
    # 1. LOAD ANALYTICS
    # --------------------------------------------------
    sales = await compute_sales_analytics(period=period, token=token)
    stock = await compute_stock_analytics(period=period, token=token)

    include_stock = report_type in ("combined", "stock")

    # --------------------------------------------------
    # 2. GENERATE CONTENT
    # --------------------------------------------------
    if fmt == "pdf":
        content = _generate_sales_pdf(
            sales=sales,
            stock=stock if include_stock else None,
        )
        filename = f"analytics_{report_type}_{period.value}_{_now_label()}.pdf"
        media_type = "application/pdf"

    elif fmt == "excel":
        content = _generate_sales_excel(
            sales=sales,
            stock=stock if include_stock else None,
        )
        filename = f"analytics_{report_type}_{period.value}_{_now_label()}.xlsx"
        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        raise HTTPException(400, f"Unsupported report format: {fmt}")

    # --------------------------------------------------
    # 3. SAVE OR RETURN
    # --------------------------------------------------
    if save_to_disk:
        file_path = _save_report_to_disk(content, filename)

        return {
            "status": "success",
            "file_path": file_path,
            "filename": filename,
            "generated_at": datetime.utcnow().isoformat(),
        }

    return content, filename, media_type